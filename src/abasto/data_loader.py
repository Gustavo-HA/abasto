"""Carga eficiente del dataset principal usando Polars lazy/streaming."""

from pathlib import Path
import polars as pl
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent.parent / "data"
VENTAS_PATH = DATA_DIR / "00_Datos_Modelar.txt"
EXCEL_PATH = DATA_DIR / "CIMAT_BaseDatos.xlsx"


def scan_ventas() -> pl.LazyFrame:
    """Retorna un LazyFrame (sin cargar en memoria) del histórico de ventas."""
    return pl.scan_csv(
        VENTAS_PATH,
        separator="|",
        schema_overrides={
            "Loc": pl.Int32,
            "Sku": pl.Int32,
            "Fecha": pl.Date,
            "Uni": pl.Int32,
        },
    ).rename({
        "Loc": "tienda_id",
        "Sku": "articulo_id",
        "Fecha": "fecha",
        "Uni": "unidades_vendidas",
    })


def load_catalogo() -> dict[str, pl.DataFrame]:
    """Carga todas las hojas del Excel (archivo pequeño, cabe en RAM)."""
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    return {name: pl.read_excel(str(EXCEL_PATH), sheet_name=name) for name in wb.sheetnames}
